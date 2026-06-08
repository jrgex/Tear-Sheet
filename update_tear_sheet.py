#!/usr/bin/env python3
"""
BEATS Tear Sheet Generator
Reads the source data sheet, computes performance statistics,
and writes a styled Excel tear sheet matching the PDF layout.

Supported source formats:
  1. Tear Sheet xlsx  – 'Table 1' sheet, year×month grid (values in %)
  2. Beats data xlsx  – two-column list: Month | decimal return
"""

import sys
import math
from pathlib import Path

import numpy as np
import pandas as pd
import xlsxwriter

BENCHMARK_TICKER  = "EWA"
BENCHMARK_NAME    = "iShares MSCI Australia ETF"

# ─── Paths ────────────────────────────────────────────────────────────────────
ROOT       = Path(__file__).parent
OUTPUT_DIR = ROOT / "output"
LOGO_FILE  = ROOT / "assets" / "gex_logo.png"

# All files in data/ are merged in order — later files override earlier ones
# for any overlapping months, so keep the most-current file last.
DATA_FILES = [
    ROOT / "data" / "Tear Sheet 14032026.xlsx",   # Jul 2022 → Feb 2026
    ROOT / "data" / "Tear_Sheet_Apr2026.xlsx",    # Jan 2023 → Apr 2026 (updated)
]

# ─── Fund constants ───────────────────────────────────────────────────────────
FUND_NAME_1     = "GREEN RENEWABLE STORAGE ENERGY MARKET"
FUND_NAME_2     = "PARTICIPATION AND TRADING"
FUND_NAME_SHORT = "Green Renewable Storage Energy Market Participation and Trading"
MANAGER         = "Green Exchange"
MIN_INVESTMENT  = "30,000 AUD"
LIQUIDITY       = "Quarterly"
MGMT_FEE        = "2.00%"
PERF_FEE        = "15.00%"
HWM             = "Yes"
PHONE           = "+61 2 8004 0234"
EMAIL           = "jr@greenexchange.com.au"

LEGAL_NOTE = (
    "NOTE: This publication has been prepared on behalf of and issued by Green Exchange "
    "Pty Ltd (ACN 654 125 247) holding AFSL 559619. This is for educational purposes only. "
    "This is not an offer to deal in any financial product. This information might contain "
    "unsolicited general information only, without regard to any investor's individual "
    "objectives, financial situation or needs. It is not specific advice for any particular "
    "investor. Before making any decision about the information provided, you must consider "
    "the appropriateness of the information in this document, having regard to your objectives, "
    "financial situation and needs and consult your adviser. This may not be passed on to anyone "
    "but the addressee. Past performance of financial products is no assurance of future performance."
)

STRATEGY_LINES = [
    "BEATS  Battery Electronic Automated Trading System",
    "",
    "Deployment of strategic Battery assets across Eastern Australia. The Energy landscape",
    "in Australia presents an Investment opportunity in BESS. This offers multiple revenue",
    "streams across the entire value chain. Installation of hardware with a 10+ year lifespan",
    "and market participation through proprietary software enable IRR of 26%+ with a payback",
    "period of 2-3 years.",
    "",
    "Certificate Generation and Trading coupled with creative commercial models further add",
    "to yields.",
]

KEY_HIGHLIGHTS = [
    "Battery System Installation (BESS)",
    "System Integration",
    "Certificate Generation",
    "Trading",
    "Commercial Models",
]

MONTH_NAMES = ["Jan","Feb","Mar","Apr","May","Jun","Jul","Aug","Sep","Oct","Nov","Dec"]

# ─── Colours ──────────────────────────────────────────────────────────────────
C_GREEN        = "#2E7D32"
C_GREEN_DARK   = "#1B5E20"
C_BLACK        = "#1A1A1A"
C_GREY         = "#757575"
C_LGREY        = "#F5F5F5"
C_WHITE        = "#FFFFFF"
C_RED          = "#C62828"
C_BORDER_LIGHT = "#E0E0E0"

# ─── Column layout (16 content cols + 2 margin cols = 18 total) ───────────────
#  Col 0  : left margin
#  Cols 1–8  (8 cols): LEFT PANEL
#  Cols 9–15 (7 cols): RIGHT PANEL
#      Right label area : cols 9–13
#      Right value area : cols 14–15
#  Col 16 : right margin (used for table overflow)
#
#  Monthly table reuses same cols 1–15 (Year=1, Jan–Dec=2–13, Annual=14–15)

MARGIN_L = 0
LS, LE   = 1, 8      # left start / end
RS, RE   = 9, 15     # right start / end
RL_END   = 13        # right-label end (merged R label: RS:RL_END)
RV_S     = 14        # right-value start (RV_S:RE)
MARGIN_R = 16

# Table column mapping (logical 0–13 → worksheet col)
# 0=Year(col1), 1–12=Jan–Dec(cols2–13), 13=Annual(cols14–15 merged)
T_YEAR   = LS           # col 1
T_JAN    = LS + 1       # col 2
T_DEC    = LS + 12      # col 13
T_ANN_S  = LS + 13      # col 14
T_ANN_E  = LS + 14      # col 15


# ═══════════════════════════════════════════════════════════════════════════════
#  Data loading
# ═══════════════════════════════════════════════════════════════════════════════

def load_data(filepaths: list[Path]) -> pd.DataFrame:
    """Load and merge returns from multiple source files.

    Files are processed in order; later files override earlier ones for any
    overlapping months, so list the most current file last.
    """
    import openpyxl
    frames = []
    for fp in filepaths:
        if not fp.exists():
            print(f"  (skipping missing file: {fp.name})")
            continue
        wb  = openpyxl.load_workbook(str(fp), data_only=True)
        df  = _load_workbook(wb, fp.name)
        frames.append(df)
        print(f"  {fp.name}: {len(df)} rows "
              f"({df['month'].iloc[0].strftime('%b %Y')} → "
              f"{df['month'].iloc[-1].strftime('%b %Y')})")

    if not frames:
        raise FileNotFoundError("No data files found in data/")

    combined = pd.concat(frames, ignore_index=True).sort_values("month")
    # Later rows override earlier ones for the same month
    combined = combined.drop_duplicates(subset=["month"], keep="last")
    return combined.sort_values("month").reset_index(drop=True)


def _load_workbook(wb, filename: str) -> pd.DataFrame:
    """Dispatch to the right loader based on sheet names present."""
    if "Table 1" in wb.sheetnames:
        return _load_tear_sheet_format(wb["Table 1"])
    if "Data" in wb.sheetnames:
        return _load_data_sheet(wb["Data"])
    if "Tear Sheet" in wb.sheetnames:
        return _load_output_sheet(wb["Tear Sheet"])
    return _load_list_format(wb.active)


def _load_data_sheet(ws) -> pd.DataFrame:
    """Read from the _data sheet generated by this script.

    Columns: Date(0), VAMI(1), _, Date(3), Return%(4)
    """
    rows = list(ws.iter_rows(min_row=2, values_only=True))  # skip header
    records = []
    for row in rows:
        if len(row) < 5 or row[3] is None or row[4] is None:
            continue
        try:
            records.append({
                "month":  pd.Timestamp(row[3]),
                "return": float(row[4]) / 100.0,
            })
        except (TypeError, ValueError):
            pass
    return pd.DataFrame(records).drop_duplicates("month").sort_values("month").reset_index(drop=True)


def _load_output_sheet(ws) -> pd.DataFrame:
    """Read the year×month table from a 'Tear Sheet' output sheet."""
    rows = list(ws.iter_rows(values_only=True))
    # Find the header row containing 'Jan'
    hdr_row = next(
        (i for i, r in enumerate(rows)
         if any(str(v).strip().lower() == "jan" for v in r if v is not None)),
        None,
    )
    if hdr_row is None:
        return pd.DataFrame(columns=["month", "return"])

    month_abbrs = ["jan","feb","mar","apr","may","jun","jul","aug","sep","oct","nov","dec"]
    col_to_month: dict[int, int] = {}
    for ci, val in enumerate(rows[hdr_row]):
        if val is None:
            continue
        s = str(val).strip().lower()
        for mi, abbr in enumerate(month_abbrs):
            if s == abbr:
                col_to_month[ci] = mi + 1

    records = []
    for row in rows[hdr_row + 1:]:
        if not row or row[0] is None:
            continue
        try:
            year = int(row[0])
        except (TypeError, ValueError):
            continue
        if not (2000 <= year <= 2100):
            continue
        for ci, month_num in col_to_month.items():
            val = row[ci] if ci < len(row) else None
            if val is None:
                continue
            try:
                records.append({
                    "month":  pd.Timestamp(year=year, month=month_num, day=1),
                    "return": float(val) / 100.0,
                })
            except (TypeError, ValueError):
                pass
    return pd.DataFrame(records).sort_values("month").reset_index(drop=True)


def _load_tear_sheet_format(ws) -> pd.DataFrame:
    rows = list(ws.iter_rows(values_only=True))
    perf_row = next(
        (i for i, r in enumerate(rows)
         if any(isinstance(v, str) and "MONTHLY PERFORMANCE" in v for v in r if v)),
        None,
    )
    if perf_row is None:
        raise ValueError("Cannot find 'MONTHLY PERFORMANCE' section")

    hdr_row = next(
        (i for i in range(perf_row + 1, len(rows)) if any(v is not None for v in rows[i])),
        None,
    )
    month_abbrs = ["jan","feb","mar","apr","may","jun","jul","aug","sep","oct","nov","dec"]
    col_to_month: dict[int, int] = {}
    for ci, val in enumerate(rows[hdr_row]):
        if val is None:
            continue
        s = str(val).strip().lower()
        for mi, abbr in enumerate(month_abbrs):
            if s == abbr:
                col_to_month[ci] = mi + 1

    records = []
    for row in rows[hdr_row + 1:]:
        if row[0] is None:
            continue
        try:
            year = int(row[0])
        except (TypeError, ValueError):
            continue
        if not (2000 <= year <= 2100):
            continue
        for ci, month_num in col_to_month.items():
            val = row[ci] if ci < len(row) else None
            if val is None:
                continue
            try:
                records.append({
                    "month":  pd.Timestamp(year=year, month=month_num, day=1),
                    "return": float(val) / 100.0,
                })
            except (TypeError, ValueError):
                pass

    return pd.DataFrame(records).sort_values("month").reset_index(drop=True)


def _load_list_format(ws) -> pd.DataFrame:
    rows = list(ws.iter_rows(values_only=True))
    hdr = next(
        (i for i, r in enumerate(rows)
         if any(isinstance(v, str) and
                ("month" in v.lower() or "beat" in v.lower()) for v in r if v)),
        0,
    )
    records = []
    for row in rows[hdr + 1:]:
        vals = [v for v in row if v is not None]
        if len(vals) < 2:
            continue
        try:
            records.append({"month": pd.Timestamp(vals[0]), "return": float(vals[1])})
        except (TypeError, ValueError):
            pass
    return pd.DataFrame(records).sort_values("month").reset_index(drop=True)


# ═══════════════════════════════════════════════════════════════════════════════
#  Benchmark (EWA)
# ═══════════════════════════════════════════════════════════════════════════════

def fetch_benchmark(start: pd.Timestamp, end: pd.Timestamp) -> pd.DataFrame:
    """Download EWA monthly returns aligned to the fund date range."""
    try:
        import yfinance as yf
        raw = yf.download(
            BENCHMARK_TICKER,
            start=(start - pd.DateOffset(months=2)).strftime("%Y-%m-%d"),
            end=(end   + pd.DateOffset(months=2)).strftime("%Y-%m-%d"),
            interval="1mo",
            auto_adjust=True,
            progress=False,
        )
        if raw.empty:
            raise ValueError("Empty response")

        # Flatten MultiIndex if present
        close = raw["Close"]
        if isinstance(close, pd.DataFrame):
            close = close.iloc[:, 0]

        # Normalise index to first day of each month
        close.index = pd.to_datetime(close.index).to_period("M").to_timestamp()
        close = close.sort_index()

        rets = close.pct_change().dropna()
        df   = pd.DataFrame({"month": rets.index, "return": rets.values})
        df   = df[(df["month"] >= start) & (df["month"] <= end)].reset_index(drop=True)
        return df

    except Exception as exc:
        print(f"  Warning: could not fetch {BENCHMARK_TICKER}: {exc}")
        return pd.DataFrame(columns=["month", "return"])


def build_benchmark_stats(bm_df: pd.DataFrame, fund_vami_dates: list) -> dict:
    """Compute benchmark VAMI (starting at 1000) aligned to fund dates."""
    if bm_df.empty:
        return {"vami": [], "annual_rets": {}, "monthly_table": {}}

    bm_by_month = {row["month"]: row["return"] for _, row in bm_df.iterrows()}

    # Build VAMI aligned to fund's date axis
    vami = [1000.0]
    for date in fund_vami_dates[1:]:        # fund_vami_dates[0] is the starting point
        ret = bm_by_month.get(date)
        vami.append(vami[-1] * (1 + ret) if ret is not None else None)

    # Annual returns
    monthly_table: dict[int, dict] = {}
    for _, row in bm_df.iterrows():
        y, m = row["month"].year, row["month"].month
        monthly_table.setdefault(y, {})[m] = row["return"] * 100

    annual_rets = {}
    for y, months in monthly_table.items():
        r_list = [months[m] / 100 for m in sorted(months)]
        annual_rets[y] = (np.prod([1 + r for r in r_list]) - 1) * 100

    return {"vami": vami, "annual_rets": annual_rets, "monthly_table": monthly_table}


# ═══════════════════════════════════════════════════════════════════════════════
#  Statistics
# ═══════════════════════════════════════════════════════════════════════════════

def build_stats(df: pd.DataFrame) -> dict:
    rets   = df["return"].values
    n      = len(rets)
    latest = df["month"].iloc[-1]

    total  = float(np.prod(1 + rets) - 1)
    ytd    = float(np.prod(1 + df.loc[df["month"].dt.year == latest.year, "return"].values) - 1)
    mo3    = float(np.prod(1 + rets[-3:]) - 1) if n >= 3 else total
    mo6    = float(np.prod(1 + rets[-6:]) - 1) if n >= 6 else total

    ann_ret = (1 + total) ** (12 / n) - 1
    ann_std = float(np.std(rets, ddof=1)) * math.sqrt(12)
    sharpe  = ann_ret / ann_std if ann_std else 0.0
    winning = float(np.sum(rets > 0) / n * 100)

    vami = [1000.0]
    for r in rets:
        vami.append(vami[-1] * (1 + r))

    monthly_table: dict[int, dict] = {}
    for _, row in df.iterrows():
        y, m = row["month"].year, row["month"].month
        monthly_table.setdefault(y, {})[m] = row["return"] * 100

    annual_rets = {
        y: (np.prod([1 + v/100 for v in months.values()]) - 1) * 100
        for y, months in monthly_table.items()
    }

    return {
        "total_return":  total * 100,
        "ytd":           ytd * 100,
        "three_mo_ror":  mo3 * 100,
        "six_mo_ror":    mo6 * 100,
        "sharpe":        sharpe,
        "winning_pct":   winning,
        "ann_return":    ann_ret * 100,
        "vami":          vami,
        "vami_dates":    [df["month"].iloc[0]] + df["month"].tolist(),
        "monthly_table": monthly_table,
        "annual_rets":   annual_rets,
        "latest_date":   latest,
        "all_returns":   rets.tolist(),
        "all_dates":     df["month"].tolist(),
    }


# ═══════════════════════════════════════════════════════════════════════════════
#  Excel generation
# ═══════════════════════════════════════════════════════════════════════════════

def write_excel(df: pd.DataFrame, stats: dict, output_path: Path,
                bm_stats: dict | None = None):
    wb = xlsxwriter.Workbook(str(output_path), {"nan_inf_to_errors": True})

    bm = bm_stats or {}

    ws = wb.add_worksheet("Tear Sheet")   # tab 1 — visible first
    ds = wb.add_worksheet("Data")         # tab 2 — chart source data
    ws.activate()                          # open on Tear Sheet by default

    _write_chart_data(wb, ds, df, stats, bm)
    ws.hide_gridlines(2)
    ws.set_paper(9)        # A4
    ws.set_portrait()
    ws.set_margins(left=0.5, right=0.5, top=0.45, bottom=0.45)
    ws.fit_to_pages(1, 0)
    ws.set_zoom(90)

    # ── Column widths ──────────────────────────────────────────────────────────
    ws.set_column(MARGIN_L, MARGIN_L, 0.8)   # left margin
    # Left panel (cols 1-8): total ~54 chars → ~3.6"
    ws.set_column(LS, LS,   7.0)   # col 1  – also table Year col
    ws.set_column(2,  2,    7.0)   # col 2  – also Jan
    ws.set_column(3,  3,    6.5)   # col 3  – Feb
    ws.set_column(4,  4,    6.5)   # col 4  – Mar
    ws.set_column(5,  5,    6.5)   # col 5  – Apr
    ws.set_column(6,  6,    6.5)   # col 6  – May
    ws.set_column(7,  7,    6.5)   # col 7  – Jun
    ws.set_column(LE, LE,   6.5)   # col 8  – Jul
    # Right panel (cols 9-15): total ~47 chars → ~3.1"
    ws.set_column(RS,    RS,    6.5)   # col 9  – Aug
    ws.set_column(10,    10,    6.5)   # col 10 – Sep
    ws.set_column(11,    11,    6.5)   # col 11 – Oct
    ws.set_column(12,    12,    6.5)   # col 12 – Nov
    ws.set_column(13,    13,    6.5)   # col 13 – Dec
    ws.set_column(RV_S,  RV_S,  7.0)   # col 14 – Annual start / right values
    ws.set_column(RE,    RE,    5.5)   # col 15 – Annual end / right margin
    ws.set_column(MARGIN_R, MARGIN_R, 0.8)  # right margin

    fmt = _build_formats(wb)
    s   = stats
    date_str = s["latest_date"].strftime("%B %Y")

    def L(r, text, f):
        ws.merge_range(r, LS, r, LE, text, f)

    def R(r, text, f):
        ws.merge_range(r, RS, r, RE, text, f)

    def RL(r, text, f):
        """Right label area (RS:RL_END)."""
        ws.merge_range(r, RS, r, RL_END, text, f)

    def RV(r, text, f):
        """Right value area (RV_S:RE)."""
        ws.merge_range(r, RV_S, r, RE, text, f)

    def FULL(r, text, f):
        ws.merge_range(r, LS, r, RE, text, f)

    row = 0

    # ── DATE + LOGO ───────────────────────────────────────────────────────────
    # Rows 0-2 form the header area; logo spans all three rows on the right.
    ws.set_row(row, 14)
    L(row, date_str, fmt["date_line"])
    R(row, "", fmt["body_r"])
    row += 1

    # ── FUND NAME ─────────────────────────────────────────────────────────────
    ws.set_row(row, 26)
    FULL(row, FUND_NAME_1, fmt["fund_name"])
    row += 1
    ws.set_row(row, 26)
    FULL(row, FUND_NAME_2, fmt["fund_name"])
    row += 1

    # Insert GEX logo top-right — exact target: 2.27 cm wide × 2.15 cm tall
    # Logo is 720×693 px at 288 DPI.  Natural display = w/dpi*2.54 cm
    # x_scale = target_cm * dpi / (px * 2.54)
    if LOGO_FILE.exists():
        IMG_DPI = 288
        logo_x_scale = 2.27 * IMG_DPI / (720 * 2.54)   # = 0.3576
        logo_y_scale = 2.15 * IMG_DPI / (693 * 2.54)   # = 0.3518
        ws.insert_image(
            0, RE,
            str(LOGO_FILE),
            {
                "x_scale":   logo_x_scale,
                "y_scale":   logo_y_scale,
                "x_offset":  -82,
                "y_offset":  2,
                "positioning": 1,
            },
        )

    # ── DIVIDER ───────────────────────────────────────────────────────────────
    ws.set_row(row, 4)
    ws.merge_range(row, MARGIN_L, row, MARGIN_R, "", fmt["divider_bar"])
    row += 1

    ws.set_row(row, 8)   # spacer
    row += 1

    # ── KEY HIGHLIGHTS | MANAGER ──────────────────────────────────────────────
    ws.set_row(row, 17)
    L(row, "KEY HIGHLIGHTS", fmt["sec_hdr_l"])
    R(row, "MANAGER",        fmt["sec_hdr_r"])
    row += 1

    for i, kh in enumerate(KEY_HIGHLIGHTS):
        ws.set_row(row, 15)
        L(row, f"  •  {kh}", fmt["bullet"])
        if i == 0:
            R(row, "Green Exchange", fmt["manager_name"])
        else:
            R(row, "", fmt["body_r"])
        row += 1

    # ── STRATEGY / PERFORMANCE STATISTICS ────────────────────────────────────
    ws.set_row(row, 10)  # spacer
    L(row, "", fmt["body_l"]); R(row, "", fmt["body_r"])
    row += 1

    ws.set_row(row, 17)
    L(row, "STRATEGY DESCRIPTION",   fmt["sec_hdr_l"])
    R(row, "PERFORMANCE STATISTICS", fmt["sec_hdr_r"])
    row += 1

    # Stat row 1: labels
    ws.set_row(row, 12)
    L(row, STRATEGY_LINES[0], fmt["strat_title"])
    RL(row, "3 Month ROR",  fmt["stat_lbl"])
    RV(row, "Year To Date", fmt["stat_lbl"])
    row += 1

    # Stat row 2: values
    ws.set_row(row, 18)
    L(row, "", fmt["body_l"])
    RL(row, f"{s['three_mo_ror']:.2f}%", fmt["stat_num"])
    RV(row, f"{s['ytd']:.2f}%",          fmt["stat_num"])
    row += 1

    # Stat row 3: labels
    ws.set_row(row, 12)
    L(row, STRATEGY_LINES[2], fmt["body_s"])
    RL(row, "Total Return Cumulative", fmt["stat_lbl"])
    RV(row, "6 Month ROR",             fmt["stat_lbl"])
    row += 1

    # Stat row 4: values
    ws.set_row(row, 18)
    L(row, STRATEGY_LINES[3], fmt["body_s"])
    RL(row, f"{s['total_return']:.2f}%", fmt["stat_num_lg"])
    RV(row, f"{s['six_mo_ror']:.2f}%",   fmt["stat_num"])
    row += 1

    # ── STRATEGY (cont.) / GENERAL INFORMATION ────────────────────────────────
    ws.set_row(row, 10)  # spacer
    L(row, STRATEGY_LINES[4], fmt["body_s"]); R(row, "", fmt["body_r"])
    row += 1

    ws.set_row(row, 17)
    L(row, STRATEGY_LINES[5], fmt["body_s"])
    R(row, "GENERAL INFORMATION", fmt["sec_hdr_r"])
    row += 1

    gi_rows = [
        ("Minimum Investment", MIN_INVESTMENT),
        ("Liquidity",          LIQUIDITY),
        ("Management Fee",     MGMT_FEE),
        ("Performance Fee",    PERF_FEE),
        ("Highwater Mark",     HWM),
    ]
    strat_rest = STRATEGY_LINES[6:]
    for i, (label, val) in enumerate(gi_rows):
        ws.set_row(row, 15)
        sl = strat_rest[i] if i < len(strat_rest) else ""
        L(row, sl, fmt["body_s"])
        RL(row, f"  {label}", fmt["gi_lbl"])
        RV(row, val,          fmt["gi_val"])
        row += 1

    # ── PERFORMANCE (VAMI) | STATISTICS ───────────────────────────────────────
    ws.set_row(row, 10)  # spacer
    L(row, "", fmt["body_l"]); R(row, "", fmt["body_r"])
    row += 1

    ws.set_row(row, 17)
    L(row, "PERFORMANCE (VAMI)", fmt["sec_hdr_l"])
    R(row, "STATISTICS",         fmt["sec_hdr_r"])
    vami_anchor = row + 1
    row += 1

    stat_items = [
        ("Total Return Cumulative", f"{s['total_return']:.2f}%"),
        ("Sharpe Ratio",            f"{s['sharpe']:.2f}"),
        ("Winning Months (%)",      f"{s['winning_pct']:.2f}%"),
        ("Alpha Annualized",        f"{s['ann_return']:.2f}%"),
    ]
    for label, val in stat_items:
        ws.set_row(row, 15)
        L(row, "", fmt["body_l"])
        RL(row, f"  {label}", fmt["stats_lbl"])
        RV(row, val,          fmt["stats_val"])
        row += 1

    vami_rows = 13   # chart height in rows
    chart_end = vami_anchor + vami_rows
    while row < chart_end:
        ws.set_row(row, 15)
        L(row, "", fmt["body_l"]); R(row, "", fmt["body_r"])
        row += 1

    # Embed VAMI chart (left panel only)
    vami_chart = _make_vami_chart(wb, ds, stats, bm)
    ws.insert_chart(vami_anchor, LS, vami_chart,
                    {"x_offset": 0, "y_offset": 2, "x_scale": 1.0, "y_scale": 1.0})

    # ── MONTHLY RETURNS CHART ─────────────────────────────────────────────────
    ws.set_row(row, 10)  # spacer
    row += 1

    ws.set_row(row, 17)
    FULL(row, "MONTHLY RETURNS", fmt["sec_hdr_full"])
    mr_anchor = row + 1
    row += 1

    mr_rows = 11
    for _ in range(mr_rows):
        ws.set_row(row, 15)
        FULL(row, "", fmt["body_l"])
        row += 1

    mr_chart = _make_monthly_returns_chart(wb, ds, stats, bm)
    ws.insert_chart(mr_anchor, LS, mr_chart,
                    {"x_offset": 0, "y_offset": 2, "x_scale": 1.0, "y_scale": 1.0})

    # ── MONTHLY PERFORMANCE TABLE ─────────────────────────────────────────────
    ws.set_row(row, 8)   # spacer
    row += 1

    ws.set_row(row, 17)
    FULL(row, "MONTHLY PERFORMANCE", fmt["sec_hdr_full"])
    row += 1

    # Table header row
    has_bm = bool(bm.get("annual_rets"))
    ws.set_row(row, 16)
    ws.write(row, T_YEAR, "", fmt["th"])
    for mi, mn in enumerate(MONTH_NAMES):
        ws.write(row, T_JAN + mi, mn, fmt["th"])
    ws.write(row, T_ANN_S, "Year Return", fmt["th"])
    ws.write(row, T_ANN_E, BENCHMARK_TICKER if has_bm else "", fmt["th"])
    row += 1

    # Table data rows
    mt = stats["monthly_table"]
    ar = stats["annual_rets"]
    bm_ar = bm.get("annual_rets", {})
    for yi, year in enumerate(sorted(mt.keys(), reverse=True)):
        ws.set_row(row, 14)
        alt = yi % 2 == 1
        ws.write(row, T_YEAR, year, fmt["ty_alt" if alt else "ty"])
        for m_idx in range(1, 13):
            val = mt[year].get(m_idx)
            c   = T_JAN + (m_idx - 1)
            if val is not None:
                ws.write(row, c, round(val, 2), _rf(wb, val, alt))
            else:
                ws.write_blank(row, c, None, fmt["te_alt" if alt else "te"])
        ann = ar.get(year)
        ws.write(row, T_ANN_S,
                 round(ann, 2) if ann is not None else "",
                 _rf(wb, ann, alt, bold=True) if ann is not None
                 else fmt["te_alt" if alt else "te"])
        bm_ann = bm_ar.get(year)
        ws.write(row, T_ANN_E,
                 round(bm_ann, 2) if bm_ann is not None else "",
                 _rf(wb, bm_ann, alt) if bm_ann is not None
                 else fmt["te_alt" if alt else "te"])
        row += 1

    # ── FOOTER ────────────────────────────────────────────────────────────────
    ws.set_row(row, 6)
    row += 1
    ws.set_row(row, 40)
    FULL(row, LEGAL_NOTE, fmt["legal"])
    row += 1
    ws.set_row(row, 16)
    FULL(row, f"  {MANAGER}   ·   Phone: {PHONE}   ·   {EMAIL}", fmt["contact"])

    wb.close()
    print(f"✓  Saved: {output_path}")


# ─── Chart data ───────────────────────────────────────────────────────────────

def _write_chart_data(wb, ds, df: pd.DataFrame, stats: dict, bm: dict):
    # Fund VAMI (cols 0-1)
    ds.write(0, 0, "Date");  ds.write(0, 1, FUND_NAME_SHORT)
    for i, (d, v) in enumerate(zip(stats["vami_dates"], stats["vami"]), 1):
        ds.write(i, 0, d.strftime("%b %Y") if hasattr(d, "strftime") else str(d))
        ds.write(i, 1, round(v, 2))

    # Fund monthly returns (cols 3-4)
    ds.write(0, 3, "Date");  ds.write(0, 4, "Fund Return (%)")
    for i, row in df.iterrows():
        ds.write(i + 1, 3, row["month"].strftime("%b %Y"))
        ds.write(i + 1, 4, round(row["return"] * 100, 4))

    # Benchmark VAMI (cols 6-7) — same date labels as fund (col 0)
    if bm.get("vami"):
        ds.write(0, 6, "Date");  ds.write(0, 7, BENCHMARK_NAME)
        for i, v in enumerate(bm["vami"], 1):
            if v is not None:
                ds.write(i, 6, stats["vami_dates"][i - 1].strftime("%b %Y")
                         if hasattr(stats["vami_dates"][i - 1], "strftime") else "")
                ds.write(i, 7, round(v, 2))

    # Benchmark monthly returns (cols 9-10) — aligned to fund monthly dates
    if bm.get("monthly_table"):
        bm_by_month = {
            pd.Timestamp(year=y, month=m, day=1): pct
            for y, months in bm["monthly_table"].items()
            for m, pct in months.items()
        }
        ds.write(0, 9, "Date");  ds.write(0, 10, BENCHMARK_NAME)
        for i, row in df.iterrows():
            val = bm_by_month.get(row["month"])
            ds.write(i + 1, 9,  row["month"].strftime("%b %Y"))
            ds.write(i + 1, 10, round(val, 4) if val is not None else "")


def _make_vami_chart(wb, ds, stats: dict, bm: dict):
    n = len(stats["vami"])
    chart = wb.add_chart({"type": "line"})
    # Fund series
    chart.add_series({
        "name":       FUND_NAME_SHORT,
        "categories": ["Data", 1, 0, n, 0],
        "values":     ["Data", 1, 1, n, 1],
        "line":       {"color": C_GREEN, "width": 2.0},
        "marker":     {"type": "none"},
    })
    # Benchmark series
    if bm.get("vami"):
        nb = len(bm["vami"])
        chart.add_series({
            "name":       BENCHMARK_NAME,
            "categories": ["Data", 1, 6, nb, 6],
            "values":     ["Data", 1, 7, nb, 7],
            "line":       {"color": "#9E9E9E", "width": 1.25, "dash_type": "dash"},
            "marker":     {"type": "none"},
        })
    chart.set_title({"none": True})
    chart.set_x_axis({
        "name": "", "num_font": {"size": 7, "color": C_GREY},
        "line": {"color": C_BORDER_LIGHT},
        "major_gridlines": {"visible": False},
    })
    chart.set_y_axis({
        "name": "Performance (VAMI)",
        "name_font": {"size": 7, "color": C_GREY},
        "num_font":  {"size": 7, "color": C_GREY},
        "line":      {"color": C_BORDER_LIGHT},
        "major_gridlines": {"visible": True,
                            "line": {"color": "#EEEEEE", "dash_type": "dash"}},
        "num_format": "#,##0",
    })
    chart.set_legend({"position": "bottom", "font": {"size": 7, "color": C_GREY}})
    chart.set_chartarea({"border": {"color": C_BORDER_LIGHT}, "fill": {"color": C_WHITE}})
    chart.set_plotarea({"fill": {"color": C_WHITE}})
    chart.set_size({"width": 385, "height": 248})
    return chart


def _make_monthly_returns_chart(wb, ds, stats: dict, bm: dict):
    n = len(stats["all_returns"])

    # Column chart — fund monthly returns
    col_chart = wb.add_chart({"type": "column"})
    col_chart.add_series({
        "name":       FUND_NAME_SHORT,
        "categories": ["Data", 1, 3, n, 3],
        "values":     ["Data", 1, 4, n, 4],
        "fill":       {"color": C_GREEN},
        "border":     {"color": C_GREEN},
        "gap":        40,
    })
    col_chart.set_title({"none": True})
    col_chart.set_x_axis({
        "name": "", "num_font": {"size": 6, "color": C_GREY},
        "line": {"color": C_BORDER_LIGHT},
        "major_gridlines": {"visible": False},
    })
    col_chart.set_y_axis({
        "name": "Monthly Return (%)",
        "name_font": {"size": 7, "color": C_GREY},
        "num_font":  {"size": 7, "color": C_GREY},
        "line":      {"color": C_BORDER_LIGHT},
        "major_gridlines": {"visible": True,
                            "line": {"color": "#EEEEEE", "dash_type": "dash"}},
    })
    col_chart.set_legend({"position": "bottom", "font": {"size": 7, "color": C_GREY}})
    col_chart.set_chartarea({"border": {"color": C_BORDER_LIGHT}, "fill": {"color": C_WHITE}})
    col_chart.set_plotarea({"fill": {"color": C_WHITE}})
    col_chart.set_size({"width": 755, "height": 210})

    # Line overlay — EWA benchmark monthly returns
    if bm.get("monthly_table"):
        line_chart = wb.add_chart({"type": "line"})
        line_chart.add_series({
            "name":       BENCHMARK_NAME,
            "categories": ["Data", 1, 9, n, 9],
            "values":     ["Data", 1, 10, n, 10],
            "line":       {"color": "#9E9E9E", "width": 1.5, "dash_type": "dash"},
            "marker":     {"type": "none"},
        })
        col_chart.combine(line_chart)

    return col_chart


# ─── Formats ──────────────────────────────────────────────────────────────────

def _build_formats(wb) -> dict:
    def mk(**kw):
        return wb.add_format({"font_name": "Calibri", "bg_color": C_WHITE, **kw})

    return {
        "date_line":   mk(font_size=9,  font_color=C_GREY,  align="left",  valign="vcenter"),
        "fund_name":   mk(font_size=18, font_color=C_GREEN, bold=True, align="left",  valign="vcenter"),
        "divider_bar": mk(bg_color=C_GREEN, font_color=C_GREEN),

        "sec_hdr_l":   mk(font_size=10, font_color=C_GREEN, bold=True,
                           align="left", valign="vcenter",
                           bottom=2, border_color=C_GREEN),
        "sec_hdr_r":   mk(font_size=10, font_color=C_GREEN, bold=True,
                           align="left", valign="vcenter",
                           bottom=2, border_color=C_GREEN),
        "sec_hdr_full": mk(font_size=10, font_color=C_GREEN, bold=True,
                           align="left", valign="vcenter",
                           bottom=2, border_color=C_GREEN),

        "bullet":      mk(font_size=9,  font_color=C_BLACK, align="left", valign="vcenter"),
        "strat_title": mk(font_size=9,  font_color=C_BLACK, bold=True, italic=True,
                           align="left", valign="vcenter"),
        "body_s":      mk(font_size=8.5, font_color=C_BLACK, align="left", valign="vcenter"),
        "body_l":      mk(font_size=9,  font_color=C_BLACK, align="left", valign="vcenter"),
        "body_r":      mk(font_size=9,  font_color=C_BLACK, align="left", valign="vcenter"),

        "manager_name": mk(font_size=11, font_color=C_BLACK, bold=True,
                           align="left", valign="vcenter"),

        "stat_lbl":    mk(font_size=8,  font_color=C_GREY,  align="left", valign="bottom"),
        "stat_num":    mk(font_size=12, font_color=C_BLACK, bold=True,
                           align="left", valign="vcenter"),
        "stat_num_lg": mk(font_size=12, font_color=C_BLACK, bold=True,
                           align="left", valign="vcenter"),

        "gi_lbl":      mk(font_size=9,  font_color=C_BLACK, align="left", valign="vcenter",
                           bottom=1, border_color=C_BORDER_LIGHT),
        "gi_val":      mk(font_size=9,  font_color=C_BLACK, bold=True,
                           align="right", valign="vcenter",
                           bottom=1, border_color=C_BORDER_LIGHT),

        "stats_lbl":   mk(font_size=9,  font_color=C_BLACK, align="left", valign="vcenter",
                           bottom=1, border_color=C_BORDER_LIGHT),
        "stats_val":   mk(font_size=9,  font_color=C_GREEN, bold=True,
                           align="right", valign="vcenter",
                           bottom=1, border_color=C_BORDER_LIGHT),

        # Table
        "th":   wb.add_format({
            "font_name": "Calibri", "font_size": 9, "bold": True,
            "font_color": C_WHITE,  "bg_color": C_GREEN_DARK,
            "align": "center", "valign": "vcenter",
            "border": 1, "border_color": C_GREEN_DARK,
        }),
        "ty":   wb.add_format({
            "font_name": "Calibri", "font_size": 9, "bold": True,
            "font_color": C_BLACK, "bg_color": C_WHITE,
            "align": "center", "valign": "vcenter",
            "border": 1, "border_color": C_BORDER_LIGHT,
        }),
        "ty_alt": wb.add_format({
            "font_name": "Calibri", "font_size": 9, "bold": True,
            "font_color": C_BLACK, "bg_color": C_LGREY,
            "align": "center", "valign": "vcenter",
            "border": 1, "border_color": C_BORDER_LIGHT,
        }),
        "te":     wb.add_format({
            "font_name": "Calibri", "font_size": 9,
            "bg_color": C_WHITE, "align": "center", "valign": "vcenter",
            "border": 1, "border_color": C_BORDER_LIGHT,
        }),
        "te_alt": wb.add_format({
            "font_name": "Calibri", "font_size": 9,
            "bg_color": C_LGREY, "align": "center", "valign": "vcenter",
            "border": 1, "border_color": C_BORDER_LIGHT,
        }),

        "legal":   mk(font_size=6.5, font_color=C_GREY, align="left", valign="top",
                       text_wrap=True, top=1, border_color=C_BORDER_LIGHT),
        "contact": mk(font_size=8.5, font_color=C_BLACK, align="left", valign="vcenter"),
    }


_rf_cache: dict = {}

def _rf(wb, value: float, alt: bool, bold: bool = False):
    """Return format for a monthly return value."""
    key = (value >= 0, alt, bold)
    if key not in _rf_cache:
        fg = C_GREEN if value >= 0 else C_RED
        bg = C_LGREY if alt else C_WHITE
        _rf_cache[key] = wb.add_format({
            "font_name": "Calibri", "font_size": 9, "bold": bold,
            "font_color": fg, "bg_color": bg,
            "align": "center", "valign": "vcenter", "num_format": "0.00",
            "border": 1, "border_color": C_BORDER_LIGHT,
        })
    return _rf_cache[key]


# ═══════════════════════════════════════════════════════════════════════════════
#  Entry point
# ═══════════════════════════════════════════════════════════════════════════════

def main():
    OUTPUT_DIR.mkdir(exist_ok=True)

    print("Loading data …")
    df = load_data(DATA_FILES)
    print(f"  Merged: {len(df)} monthly records  "
          f"({df['month'].iloc[0].strftime('%b %Y')} → "
          f"{df['month'].iloc[-1].strftime('%b %Y')})")

    stats  = build_stats(df)
    latest = stats["latest_date"]
    fname  = f"Tear_Sheet_{latest.strftime('%b%Y')}.xlsx"
    out    = OUTPUT_DIR / fname

    print(f"Fetching benchmark ({BENCHMARK_TICKER}) …")
    bm_df    = fetch_benchmark(df["month"].iloc[0], latest)
    bm_stats = build_benchmark_stats(bm_df, stats["vami_dates"])
    if bm_df.empty:
        print("  Benchmark unavailable — chart will show fund only.")
    else:
        print(f"  {BENCHMARK_TICKER}: {len(bm_df)} months fetched")

    print("Building tear sheet …")
    write_excel(df, stats, out, bm_stats)

    print()
    print("  Total Return Cumulative :", f"{stats['total_return']:.2f}%")
    print("  YTD                     :", f"{stats['ytd']:.2f}%")
    print("  3-Month ROR             :", f"{stats['three_mo_ror']:.2f}%")
    print("  6-Month ROR             :", f"{stats['six_mo_ror']:.2f}%")
    print("  Sharpe Ratio            :", f"{stats['sharpe']:.2f}")
    print("  Winning Months          :", f"{stats['winning_pct']:.2f}%")


if __name__ == "__main__":
    main()
